"""
Сервисный слой для Asset — по п. 4.2 ТЗ и сессии B06 посессионного плана.

list_assets() — листинг с фильтрами (status, type_id, location, search) и пагинацией.
get_asset() — получение одного актива с подгруженными movements/repairs.
create_asset() — создание нового актива.
"""
import uuid
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from sqlalchemy import func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from app.models.asset import Asset, AssetStatus
from app.models.movement import Movement
from app.models.ticket import Ticket, TicketStatus
from app.schemas.asset import AssetCreate, AssetUpdate


class AssetHasOpenTicketsError(Exception):
    """Поднимается при попытке перевести Asset в written_off, пока по нему есть
    открытые заявки (Ticket.status not in DONE/REJECTED). Обрабатывается в api/assets.py
    как 409 Conflict — намеренно не HTTPException здесь, сервисный слой фреймворк-агностичен."""

class AssetDuplicateError(Exception):
    """Поднимается при попытке создать/обновить Asset с уже занятым inventory_number
    (UNIQUE constraint). Обрабатывается в api/assets.py как 409 Conflict."""

def _apply_filters(
    query: Any,
    *,
    status: AssetStatus | None,
    type_id: uuid.UUID | None,
    location: str | None,
    search: str | None,
) -> Any:
    """Применяет общий набор фильтров и к count-запросу, и к запросу за данными —
    чтобы условия не расходились между total и items."""
    if status is not None:
        query = query.where(Asset.status == status)
    if type_id is not None:
        query = query.where(Asset.type_id == type_id)
    if location is not None:
        query = query.where(Asset.location == location)
    if search:
        pattern = f"%{search}%"
        query = query.where(
            or_(
                Asset.inventory_number.ilike(pattern),
                Asset.model.ilike(pattern),
                Asset.serial_number.ilike(pattern),
            )
        )
    return query


def list_assets(
    db: Session,
    *,
    status: AssetStatus | None = None,
    type_id: uuid.UUID | None = None,
    location: str | None = None,
    search: str | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[Asset], int]:
    """Возвращает (items, total) с учётом фильтров и пагинации.

    search ищет по inventory_number/model/serial_number через ILIKE (регистронезависимо).
    total считается отдельным агрегатным запросом (SELECT count(*)), без подгрузки строк —
    не тянет из БД полные объекты Asset только ради подсчёта количества.
    """
    count_query = _apply_filters(
        select(func.count()).select_from(Asset),
        status=status, type_id=type_id, location=location, search=search,
    )
    total = db.execute(count_query).scalar_one()

    query = select(Asset).options(
        selectinload(Asset.type), selectinload(Asset.responsible_user)
    )
    query = _apply_filters(
        query, status=status, type_id=type_id, location=location, search=search
    )
    query = query.order_by(Asset.inventory_number).offset((page - 1) * page_size).limit(page_size)
    items = list(db.execute(query).scalars().all())

    return items, total


def get_asset(db: Session, asset_id: uuid.UUID) -> Asset | None:
    """Возвращает актив по id с подгруженными type, responsible_user, movements, repairs."""
    query = (
        select(Asset)
        .where(Asset.id == asset_id)
        .options(
            selectinload(Asset.type),
            selectinload(Asset.responsible_user),
            selectinload(Asset.movements).selectinload(Movement.initiator),
            selectinload(Asset.repairs),
        )
    )
    return db.execute(query).scalar_one_or_none()

def create_asset(db: Session, data: AssetCreate) -> Asset:
    """Создаёт новый актив. Валидацию существования type_id оставляем БД (FK constraint).

    При дубликате inventory_number (UNIQUE) поднимает AssetDuplicateError вместо
    того, чтобы дать IntegrityError всплыть голым 500 (техдолг из B04/B05/B06, закрыт в B07).
    """
    asset = Asset(
        inventory_number=data.inventory_number,
        type_id=data.type_id,
        serial_number=data.serial_number,
        model=data.model,
        purchase_date=data.purchase_date,
        location=data.location,
    )
    db.add(asset)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise AssetDuplicateError(
            f"Актив с инвентарным номером «{data.inventory_number}» уже существует"
        ) from exc
    db.refresh(asset)
    return asset

def _has_open_tickets(db: Session, asset_id: uuid.UUID) -> bool:
    """True, если у актива есть хотя бы одна незакрытая заявка (NEW/IN_PROGRESS)."""
    query = select(func.count()).select_from(Ticket).where(
        Ticket.asset_id == asset_id,
        Ticket.status.not_in([TicketStatus.DONE, TicketStatus.REJECTED]),
    )
    return db.execute(query).scalar_one() > 0

def update_asset(
    db: Session,
    asset: Asset,
    data: AssetUpdate,
    initiator_id: uuid.UUID,
) -> Asset:
    """Частично обновляет актив (PATCH-семантика — только переданные поля).

    Если среди переданных полей есть status и/или location, и хотя бы одно из них
    реально меняет значение, и у актива на момент изменения есть непустой location —
    атомарно (в той же транзакции) создаётся запись Movement(from_location=<до>,
    to_location=<после>, initiator_id=<текущий пользователь>). Если location у актива
    не задан (None) — запись Movement не создаётся, т.к. to_location NOT NULL в БД
    (нечего фиксировать: актив "переезжает в никуда").

    Если data.status == written_off и по активу есть открытые заявки — поднимает
    AssetHasOpenTicketsError, никаких изменений в БД не производится (проверка
    выполняется до мутации asset и до db.commit()).
    """
    update_data = data.model_dump(exclude_unset=True)

    if update_data.get("status") == AssetStatus.WRITTEN_OFF and _has_open_tickets(db, asset.id):
        raise AssetHasOpenTicketsError(
            f"Нельзя списать актив {asset.inventory_number}: есть незакрытые заявки"
        )

    old_location = asset.location
    old_status = asset.status

    for field, value in update_data.items():
        setattr(asset, field, value)

    location_changed = "location" in update_data and asset.location != old_location
    status_changed = "status" in update_data and asset.status != old_status

    if (location_changed or status_changed) and asset.location is not None:
        db.add(
            Movement(
                asset_id=asset.id,
                from_location=old_location,
                to_location=asset.location,
                initiator_id=initiator_id,
            )
        )
    attempted_inventory_number = asset.inventory_number
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise AssetDuplicateError(
            f"Актив с инвентарным номером «{attempted_inventory_number}» уже существует"
        ) from exc
    db.refresh(asset)
    return asset

_ASSET_STATUS_LABELS: dict[AssetStatus, str] = {
    AssetStatus.IN_USE: "В работе",
    AssetStatus.REPAIR: "В ремонте",
    AssetStatus.WRITTEN_OFF: "Списано",
    AssetStatus.IN_STOCK: "На складе",
}


def export_assets_workbook(assets: list[Asset]) -> BytesIO:
    """Формирует .xlsx-отчёт по списку активов через openpyxl.

    Набор и порядок колонок намеренно зеркалит уже существующую реализацию
    Dev2 (frontend/src/api/assets.ts::createAssetWorkbook), чтобы при переключении
    UI инвентаризации с мокового экспорта на реальный API визуально ничего
    не изменилось для пользователя.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Инвентаризация"
    sheet.append(
        [
            "Инвентарный номер", "Тип", "Модель", "Серийный номер", "Статус",
            "Расположение", "Ответственный", "Дата приобретения", "Hostname", "IP-адрес",
        ]
    )
    for asset in assets:
        sheet.append(
            [
                asset.inventory_number,
                asset.type.name,
                asset.model or "",
                asset.serial_number or "",
                _ASSET_STATUS_LABELS[asset.status],
                asset.location or "",
                asset.responsible_user.full_name if asset.responsible_user else "",
                asset.purchase_date.isoformat() if asset.purchase_date else "",
                asset.hostname or "",
                str(asset.ip_address) if asset.ip_address else "",
            ]
        )

    last_row = sheet.max_row
    last_col_letter = sheet.cell(row=1, column=sheet.max_column).column_letter
    sheet.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    column_widths = [20, 24, 28, 22, 14, 24, 24, 18, 20, 16]
    for index, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
